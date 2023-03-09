import tkinter as tk
from tkinter.ttk import *
from tkinter.scrolledtext import ScrolledText
import openpyxl

class Column:

    def __init__(self, name, chrom_type, guard, low_MW, high_MW, pH_min, pH_max, particle_size,
                 manufacturer, carbon_load):
        self.chrom_type = chrom_type
        self.name = name
        self.guard = guard
        self.low_MW = low_MW
        self.high_MW = high_MW
        self.pH_min = pH_min
        self.pH_max = pH_max
        self.particle_size = particle_size
        self.manufacturer = manufacturer
        self.carbon_load = carbon_load

    def print(self):
        print(self.chrom_type)
        print(self.name)
        print(self.guard)
        print(self.low_MW)
        print(self.high_MW)
        print(self.pH_min)
        print(self.pH_max)
        print(self.particle_size)
        print(self.manufacturer)
        print(self.carbon_load)

columns = []

path = "HPLC_log.xlsx"

wb_obj = openpyxl.load_workbook(path, read_only=True)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
max_column = sheet_obj.max_column

for x in range(2, max_row + 1):
    chrom = sheet_obj.cell(row=x, column=1).value
    name = sheet_obj.cell(row=x, column=2).value
    g = sheet_obj.cell(row=x, column=3).value
    low_MW = sheet_obj.cell(row=x, column=4).value
    high_MW = sheet_obj.cell(row=x, column=5).value
    ph_min = sheet_obj.cell(row=x, column=6).value
    ph_max = sheet_obj.cell(row=x, column=7).value
    particle_size = sheet_obj.cell(row=x, column=8).value
    manu = sheet_obj.cell(row=x, column=9).value
    carbon_load = sheet_obj.cell(row=x, column=10).value

    HPLC_column = Column(name, chrom, g, low_MW, high_MW, ph_min, ph_max, particle_size,
                         manu, carbon_load)

    columns.append(HPLC_column)

def on_closing():
    results.destroy()
    window.destroy()

results = tk.Tk()
results.title("HPLC Columns")
results.protocol("WM_DELETE_WINDOW", on_closing)
text_area = ScrolledText(results, wrap=tk.WORD, width=40, height=120, font = ("Times New Roman", 15))
text_area.pack()
for column in columns:
    text_area.insert(tk.INSERT, "- " + column.name + "\n")

window = tk.Tk()
window.geometry("850x500")
window.title("HPLC Column Finder")
window.protocol("WM_DELETE_WINDOW", on_closing)

for i in range(13):
    tk.Grid.rowconfigure(window, i, weight=1)

tk.Grid.columnconfigure(window, 0, weight=1)

user_chrom = None
user_manu = None
user_mw = -2
user_ph = -2
user_particle_size = -1.0
user_carbon_load = -1.0

chromatography_type = Label(window, text="Chromatography Type: ")
chromatography_type.grid(column=0, row=0, sticky="nsew")

chrom_frame = Frame(window)
chrom_frame.grid(row=1, column=0, sticky="nsew")
tk.Grid.rowconfigure(chrom_frame, 0, weight=1)
button_dict = {}
option = ["SEC", "RP", "Pep Map", "AEX", "CEX"]
chrom_column = 0
for i in option:
    def button_choice(x=i):
        text_area.delete("1.0", tk.END)
        global user_chrom
        user_chrom = x
        for column in columns:
            if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
            ):
                text_area.insert(tk.INSERT, "- " + column.name + "\n")
    tk.Grid.columnconfigure(chrom_frame, chrom_column, weight=1)
    button_dict[i] = Button(chrom_frame, text=i, command=button_choice)
    button_dict[i].grid(column=chrom_column, row=0, sticky="nsew")
    chrom_column += 1

molecular_weight = Label(window, text="Molecular weight in Daltons: ")
molecular_weight.grid(column=0, row=2, sticky="nsew")

def set_mw():
    global user_mw
    user_mw = int(mw.get())
    text_area.delete("1.0", tk.END)
    for column in columns:
        if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
        ):
            text_area.insert(tk.INSERT, "- " + column.name + "\n")

def clear_mw():
    mw.delete(0, tk.END)
    global user_mw
    user_mw = -2
    text_area.delete("1.0", tk.END)
    for column in columns:
        if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
        ):
            text_area.insert(tk.INSERT, "- " + column.name + "\n")
mw_frame = Frame(window)
mw_frame.grid(row=3, column=0, sticky="nsew")
mw = Entry(mw_frame)
mw.grid(row=0, column=0, sticky="nsew")
submit_mw = Button(mw_frame, text="Submit", command=set_mw)
submit_mw.grid(row=0, column=1, sticky="nsew")
reset_mw = Button(mw_frame, text="Reset", command=clear_mw)
reset_mw.grid(row=0, column=2, sticky="nsew")

def set_ph():
    global user_ph
    user_ph = int(ph.get())
    text_area.delete("1.0", tk.END)
    for column in columns:
        if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
        ):
            text_area.insert(tk.INSERT, "- " + column.name + "\n")

def clear_ph():
    ph.delete(0, tk.END)
    global user_ph
    user_ph = -2
    text_area.delete("1.0", tk.END)
    for column in columns:
        if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
        ):
            text_area.insert(tk.INSERT, "- " + column.name + "\n")

desired_ph = Label(window, text="Desired pH: ")
desired_ph.grid(row=4, column=0, sticky="nsew")
ph_frame = Frame(window)
ph_frame.grid(row=5, column=0, sticky="nsew")
ph = Entry(ph_frame)
ph.grid(column=0, row=0, sticky="nsew")
submit_ph = Button(ph_frame, text="Submit", command=set_ph)
submit_ph.grid(row=0, column=1, sticky="nsew")
reset_ph = Button(ph_frame, text="Reset", command=clear_ph)
reset_ph.grid(row=0, column=2, sticky="nsew")

bead_size = Label(window, text="Desired particle size: ")
bead_size.grid(column=0, row=6, sticky="nsew")

bead_frame = Frame(window)
bead_frame.grid(row=7, column=0, sticky="nsew")
tk.Grid.rowconfigure(bead_frame, 0, weight=1)
tk.Grid.rowconfigure(bead_frame, 1, weight=1)
button_dict_4 = {}
option_4 = ["1.6", "1.7", "1.8", "2.2", "2.6", "2.7", "3.0", "3.5",
            "3.6", "4.0", "5.0", "6.0", "8.0", "8.6", "9.0", "10.0", "12.0",
            "13.0", "20.0"]
chrom_column = 0
new_chrom_column = 0
for i in option_4:
    def button_choice_4(x=i):
        text_area.delete("1.0", tk.END)
        global user_particle_size
        user_particle_size = float(x)
        for column in columns:
            if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
            ):
                text_area.insert(tk.INSERT, "- " + column.name + "\n")
    tk.Grid.columnconfigure(bead_frame, chrom_column, weight=1)
    button_dict_4[i] = Button(bead_frame, text=i, command=button_choice_4)
    if chrom_column <= 9:
        button_dict_4[i].grid(column=chrom_column, row=0, sticky="nsew")
    else:
        button_dict_4[i].grid(column=new_chrom_column, row=1, sticky="nsew")
        new_chrom_column += 1
    chrom_column += 1

manu = Label(window, text="Manufacturer: ")
manu.grid(column=0, row=8, sticky="nsew")

manu_frame = Frame(window)
manu_frame.grid(row=9, column=0, sticky="nsew")
tk.Grid.rowconfigure(manu_frame, 0, weight=1)
tk.Grid.rowconfigure(manu_frame, 1, weight=1)
button_dict_2 = {}
option_2 = ["ThermoFisher", "Waters", "Agilent", "Sepax", "Cytiva", "Tosoh", "YMC", "Phenomenex",
            "Inertsil", "Sigma", "VWR", "Shodex", "GL Sciences"]
chrom_column = 0
new_chrom_column = 0
for i in option_2:
    def button_choice_2(x=i):
        text_area.delete("1.0", tk.END)
        global user_manu
        user_manu = x
        for column in columns:
            if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
            ):
                text_area.insert(tk.INSERT, "- " + column.name + "\n")
    tk.Grid.columnconfigure(manu_frame, chrom_column, weight=1)
    button_dict_2[i] = Button(manu_frame, text=i, command=button_choice_2)
    if chrom_column <= 8:
        button_dict_2[i].grid(column=chrom_column, row=0, sticky="nsew")
    else:
        button_dict_2[i].grid(column=new_chrom_column, row=1, sticky="nsew")
        new_chrom_column += 1
    chrom_column += 1

carbon = Label(window, text="Carbon Load: ")
carbon.grid(column=0, row=10, sticky="nsew")

carbon_frame = Frame(window)
carbon_frame.grid(row=11, column=0, sticky="nsew")
tk.Grid.rowconfigure(carbon_frame, 0, weight=1)
tk.Grid.rowconfigure(carbon_frame, 1, weight=1)
button_dict_3 = {}
option_3 = ["3.0", "7.0", "7.5", "8.0", "8.5", "9.0", "9.2", "10.0",
            "11.0", "12.0", "15.0", "17.0", "18.0"]
chrom_column = 0
new_chrom_column = 0
for i in option_3:
    def button_choice_3(x=i):
        text_area.delete("1.0", tk.END)
        global user_carbon_load
        user_carbon_load = float(x)
        for column in columns:
            if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
            ):
                text_area.insert(tk.INSERT, "- " + column.name + "\n")
    tk.Grid.columnconfigure(carbon_frame, chrom_column, weight=1)
    button_dict_3[i] = Button(carbon_frame, text=i, command=button_choice_3)
    if chrom_column <= 9:
        button_dict_3[i].grid(column=chrom_column, row=0, sticky="nsew")
    else:
        button_dict_3[i].grid(column=new_chrom_column, row=1, sticky="nsew")
        new_chrom_column += 1
    chrom_column += 1

def reset_all():
    global user_ph
    global user_mw
    global user_manu
    global user_chrom
    global user_particle_size
    global user_carbon_load
    user_ph = -2
    user_mw = -2
    user_manu = None
    user_chrom = None
    user_particle_size = -1.0
    user_carbon_load = -1.0
    text_area.delete("1.0", tk.END)
    for column in columns:
        if (
                (user_chrom == column.chrom_type or user_chrom == None) and
                (user_manu == column.manufacturer or user_manu == None) and
                ((user_mw >= column.low_MW and user_mw <= column.high_MW) or user_mw == -2) and
                ((user_ph >= column.pH_min and user_ph <= column.pH_max) or user_ph == -2) and
                (user_particle_size == column.particle_size or user_particle_size == -1.0) and
                (user_carbon_load == column.carbon_load or user_carbon_load == -1.0)
        ):
            text_area.insert(tk.INSERT, "- " + column.name + "\n")

reset_frame = Frame(window)
reset_frame.grid(row=12, column=0, sticky="nsew")
tk.Grid.rowconfigure(reset_frame, 0, weight=1)
tk.Grid.columnconfigure(reset_frame, 0, weight=1)
reset_all = Button(reset_frame, text="Reset All", command=reset_all)
reset_all.grid(column=0, row=0, sticky="nsew", pady=20, padx=350)

window.mainloop()

